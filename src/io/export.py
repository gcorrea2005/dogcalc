"""Export analysis results to formatted Excel (.xlsx) file.

Produces 3 sheets:
  1. Displacements — all node displacements
  2. Reactions — support reactions
  3. Member Forces — internal forces along each member's length
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side


def export_results_to_excel(doc, analysis_result, filepath: str):
    """Export analysis results to a formatted Excel workbook.

    Args:
        doc: Document with nodes, members
        analysis_result: AnalysisResult from solver
        filepath: output .xlsx path
    """
    wb = openpyxl.Workbook()

    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(start_color="003388", end_color="003388", fill_type="solid")
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def style_header(ws):
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = thin

    # ── Sheet 1: Displacements ──
    ws1 = wb.active
    ws1.title = "Displacements"
    ws1.append(["Node", "Label", "X (m)", "Y (m)", "Z (m)", "DX (mm)", "DY (mm)", "DZ (mm)"])
    style_header(ws1)

    for nid, node in doc.nodes.items():
        r = analysis_result.node_results.get(nid)
        if r:
            ws1.append([
                node.label, f"({node.x:.2f}, {node.y:.2f}, {node.z:.2f})",
                node.x, node.y, node.z,
                round(r.dx * 1000, 3), round(r.dy * 1000, 3), round(r.dz * 1000, 3),
            ])

    # ── Sheet 2: Reactions ──
    ws2 = wb.create_sheet("Reactions")
    ws2.append(["Node", "Label", "Support", "Rx (kN)", "Ry (kN)", "Rz (kN)"])
    style_header(ws2)

    for nid, node in doc.nodes.items():
        if node.is_supported:
            nr = analysis_result.node_results.get(nid)
            if nr:
                ws2.append([
                    node.label, f"({node.x:.2f}, {node.y:.2f}, {node.z:.2f})",
                    node.support_type.value,
                    round(nr.rxn_fx * 1e-3, 3),  # N → kN
                    round(nr.rxn_fy * 1e-3, 3),
                    round(nr.rxn_fz * 1e-3, 3),
                ])

    # ── Sheet 3: Member Forces ──
    ws3 = wb.create_sheet("Member Forces")
    ws3.append(["Member", "Label", "Start", "End", "x/L", "Shear V (kN)", "Moment M (kN·m)"])
    style_header(ws3)

    for mid, member in doc.members.items():
        mr = analysis_result.member_results.get(mid)
        if mr and mr.segments:
            n1_label = doc.nodes[member.start_node_id].label
            n2_label = doc.nodes[member.end_node_id].label
            for seg in mr.segments:
                ws3.append([
                    member.label, f"{n1_label}→{n2_label}",
                    n1_label, n2_label,
                    f"{seg['x']:.2f}",
                    round(seg.get('shear_y', 0) * 1e-3, 3),
                    round(seg.get('moment_z', 0) * 1e-3, 3),
                ])

    # Auto-width
    for ws in (ws1, ws2, ws3):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 28)

    wb.save(filepath)
